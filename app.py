from flask import Flask, request, jsonify, send_file, render_template
import os, io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import psycopg2
import psycopg2.extras

app = Flask(__name__)

PLAYERS = ["Koy", "Kyle", "Jensen", "Austin", "Ryan", "Hayden", "Treyson"]
HANDICAPS = {
    "Koy": 14, "Kyle": 12, "Jensen": 19, "Austin": 19,
    "Ryan": 4, "Hayden": 18, "Treyson": 20
}
COURSE_PAR = {
    "Fox Hollow - White Tees": 36, "Fox Hollow - Back 9": 36,
    "Cedar Hills - Back 9": 35, "Talons Cove": 36,
    "Sleepy Ridge": 36, "The Oaks": 36, "Green Spring": 36,
}
MIN_HOLES = 9

def get_db():
    conn = psycopg2.connect(os.environ["DATABASE_URL"], sslmode="require")
    return conn

def init_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS rounds (
        id SERIAL PRIMARY KEY, date TEXT NOT NULL,
        player TEXT NOT NULL, course TEXT NOT NULL,
        holes INTEGER NOT NULL, points INTEGER NOT NULL,
        gross INTEGER, par INTEGER, alloc INTEGER,
        net INTEGER, over_par INTEGER)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS courses (
        name TEXT PRIMARY KEY, par_9 INTEGER NOT NULL)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS players (
        name TEXT PRIMARY KEY, handicap INTEGER NOT NULL)""")
    conn.commit()
    cur.close()
    conn.close()
    seed_db()

def seed_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM rounds")
    if cur.fetchone()[0] > 0:
        cur.close(); conn.close(); return
    for course, par9 in COURSE_PAR.items():
        cur.execute("INSERT INTO courses VALUES (%s,%s) ON CONFLICT DO NOTHING", (course, par9))
    for player, hcp in HANDICAPS.items():
        cur.execute("INSERT INTO players VALUES (%s,%s) ON CONFLICT DO NOTHING", (player, hcp))
    existing_rounds = [
        ("2026-04-17","Koy","Fox Hollow - White Tees",9,10,51),
        ("2026-04-17","Kyle","Fox Hollow - White Tees",9,14,46),
        ("2026-04-17","Jensen","Fox Hollow - White Tees",9,17,46),
        ("2026-04-17","Austin","Fox Hollow - White Tees",9,15,48),
        ("2026-04-22","Koy","Cedar Hills - Back 9",9,15,45),
        ("2026-04-22","Kyle","Cedar Hills - Back 9",9,15,44),
        ("2026-04-22","Jensen","Cedar Hills - Back 9",9,16,46),
        ("2026-04-22","Austin","Cedar Hills - Back 9",9,14,49),
        ("2026-04-22","Ryan","Cedar Hills - Back 9",9,11,44),
        ("2026-04-22","Hayden","Cedar Hills - Back 9",9,16,46),
        ("2026-04-22","Treyson","Cedar Hills - Back 9",9,11,51),
        ("2026-04-24","Koy","Talons Cove",18,39,84),
        ("2026-04-24","Kyle","Talons Cove",18,29,90),
        ("2026-04-24","Jensen","Talons Cove",18,33,94),
        ("2026-04-24","Austin","Talons Cove",18,28,99),
        ("2026-05-01","Koy","Sleepy Ridge",18,37,85),
        ("2026-05-01","Kyle","Sleepy Ridge",18,30,90),
        ("2026-05-01","Jensen","Sleepy Ridge",18,39,88),
        ("2026-05-01","Austin","Sleepy Ridge",18,28,100),
        ("2026-05-05","Koy","Fox Hollow - Back 9",9,22,39),
        ("2026-05-05","Kyle","Fox Hollow - Back 9",9,13,47),
        ("2026-05-05","Jensen","Fox Hollow - Back 9",9,17,46),
        ("2026-05-05","Austin","Fox Hollow - Back 9",9,19,44),
        ("2026-05-08","Koy","The Oaks",18,25,97),
        ("2026-05-08","Kyle","The Oaks",18,28,94),
        ("2026-05-08","Jensen","The Oaks",18,27,101),
        ("2026-05-08","Treyson","Green Spring",9,16,48),
    ]
    for date, player, course, holes, points, gross in existing_rounds:
        par_9 = COURSE_PAR.get(course, 36)
        par   = round(par_9 * holes / 9)
        alloc = round(HANDICAPS.get(player, 18) * holes / 18)
        cur.execute(
            "INSERT INTO rounds (date,player,course,holes,points,gross,par,alloc,net,over_par) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (date, player, course, holes, points, gross, par, alloc, gross-alloc, gross-par))
    conn.commit()
    cur.close()
    conn.close()

def get_standings():
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT name FROM players ORDER BY name")
    players = [r["name"] for r in cur.fetchall()]
    cur.execute("SELECT * FROM rounds ORDER BY date")
    rows = cur.fetchall()
    cur.close(); conn.close()
    stats = {}
    for p in players:
        pr  = [r for r in rows if r["player"] == p]
        th  = sum(r["holes"]  for r in pr)
        tp  = sum(r["points"] for r in pr)
        pph = round(tp / th, 4) if th else 0
        stats[p] = {"rounds": len(pr), "holes": th, "points": tp, "pph": pph,
                    "handicap": HANDICAPS.get(p, 18), "history": [dict(r) for r in pr]}
    ranked = sorted([(p, s["pph"]) for p, s in stats.items() if s["holes"] >= MIN_HOLES],
                    key=lambda x: -x[1])
    for i, (p, _) in enumerate(ranked):
        stats[p]["rank"] = i + 1
    return stats, players

@app.route("/")
def index():
    return render_template("index.html", players=PLAYERS)

@app.route("/api/standings")
def api_standings():
    stats, players = get_standings()
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT name FROM courses ORDER BY name")
    courses = [r["name"] for r in cur.fetchall()]
    cur.close(); conn.close()
    return jsonify({"standings": stats, "players": players, "courses": courses})

@app.route("/api/add_round", methods=["POST"])
def add_round():
    data = request.json
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT par_9 FROM courses WHERE name=%s", (data["course"],))
    row = cur.fetchone()
    par_9 = row["par_9"] if row else 36
    cur.execute("SELECT handicap FROM players WHERE name=%s", (data["player"],))
    row = cur.fetchone()
    hcp = row["handicap"] if row else 18
    holes = int(data["holes"])
    gross = int(data["gross"]) if data.get("gross") else None
    par   = round(par_9 * holes / 9)
    alloc = round(hcp * holes / 18)
    net   = gross - alloc if gross else None
    over  = gross - par   if gross else None
    cur.execute(
        "INSERT INTO rounds (date,player,course,holes,points,gross,par,alloc,net,over_par) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
        (data["date"], data["player"], data["course"], holes, int(data["points"]), gross, par, alloc, net, over))
    conn.commit()
    cur.close(); conn.close()
    return jsonify({"status": "ok"})

@app.route("/api/add_course", methods=["POST"])
def add_course():
    data = request.json
    conn = get_db()
    cur = conn.cursor()
    cur.execute("INSERT INTO courses VALUES (%s,%s) ON CONFLICT (name) DO UPDATE SET par_9=%s",
                (data["name"], int(data["par_9"]), int(data["par_9"])))
    conn.commit()
    cur.close(); conn.close()
    return jsonify({"status": "ok"})

@app.route("/api/update_handicap", methods=["POST"])
def update_handicap():
    data = request.json
    conn = get_db()
    cur = conn.cursor()
    cur.execute("INSERT INTO players VALUES (%s,%s) ON CONFLICT (name) DO UPDATE SET handicap=%s",
                (data["player"], int(data["handicap"]), int(data["handicap"])))
    conn.commit()
    cur.close(); conn.close()
    return jsonify({"status": "ok"})

@app.route("/api/export")
def export_excel():
    stats, players = get_standings()
    wb  = _build_excel(stats, players)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name="Frends_Stableford_Championship.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def _build_excel(stats, players):
    G="1A5C2A"; GM="2E7D3E"; GOLD="FFD700"; WH="FFFFFF"; GR="F2F2F2"; BC="CCCCCC"
    def bd():
        s = Side(style="thin", color=BC)
        return Border(left=s, right=s, top=s, bottom=s)
    def cell(ws, ref, val, bg=WH, fg="000000", bold=False, sz=10, center=True, fmt=None):
        c = ws[ref]; c.value = val
        c.font = Font(name="Arial", bold=bold, color=fg, size=sz)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
        c.border = bd()
        if fmt: c.number_format = fmt
    wb = Workbook(); ws = wb.active; ws.title = "Leaderboard"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    c = ws["A1"]; c.value = "FRIENDS STABLEFORD YEAR-LONG CHAMPIONSHIP"
    c.font = Font(name="Arial", bold=True, color=WH, size=14)
    c.fill = PatternFill("solid", start_color=G)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36
    for col, h in zip(["A","B","C","D","E","F","G","H"],
                      ["Rank","Player","Handicap","Rounds","Holes","Points","Pts/Hole","Form"]):
        cell(ws, f"{col}2", h, bg=GM, fg=WH, bold=True)
    for col, w in zip(["A","B","C","D","E","F","G","H"],[7,14,9,9,8,8,10,14]):
        ws.column_dimensions[col].width = w
    ranked   = sorted([(p,stats[p]["pph"]) for p in players if stats[p].get("rank")], key=lambda x:-x[1])
    unranked = [p for p in players if not stats[p].get("rank")]
    medals   = {1:"1st",2:"2nd",3:"3rd"}
    for i, p in enumerate([p for p,_ in ranked]+unranked, start=3):
        s=stats[p]; r=s.get("rank","-")
        bg=GOLD if r==1 else (GR if i%2==0 else WH)
        cell(ws,f"A{i}",f"{medals.get(r,'')} {r}" if isinstance(r,int) else r,bg=bg,bold=(r==1))
        cell(ws,f"B{i}",p,bg=bg,bold=(r==1),center=False)
        cell(ws,f"C{i}",s["handicap"],bg=bg); cell(ws,f"D{i}",s["rounds"],bg=bg)
        cell(ws,f"E{i}",s["holes"],bg=bg);   cell(ws,f"F{i}",s["points"],bg=bg)
        cell(ws,f"G{i}",round(s["pph"],3) if s["holes"]>=MIN_HOLES else f"<{MIN_HOLES}h",bg=bg,bold=(r==1),fmt="0.000")
        cell(ws,f"H{i}","Steady",bg=bg)
    return wb

init_db()

if __name__ == "__main__":
    app.run(debug=True)
